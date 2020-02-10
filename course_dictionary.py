import re

from collections import namedtuple
from openpyxl import load_workbook
Course = namedtuple('Course', 'program, designation')
CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
class schedule():
    def __init__(self, initial_state):
        self.dict = self.create_course_dict()
        self.scheduled = {0: initial_state, 1: [] , 2: [] , 3: [], 4: [], 5: [], 6: [], 7: [], 8: []}
        self.course_translation = {'Frosh': 1, 'Soph': 3, 'Junior': 5, 'Senior': 7}
        Course = namedtuple('Course', 'program, designation')
        CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
        print(self.dict)

    def get_prereqs(self, course):
        return self.dict.get(course).prereqs

    def schedule_course(self, course, year, semester):
        if semester == 'Fall':
            self.scheduled[self.course_translation[year]].append(course)
        else:
            self.scheduled[self.course_translation[year] + 1].append(course)





    def create_course_dict(self):
        """
        Creates a dictionary containing course info.
        Keys: namedtuple of the form ('program, designation')
        Values: namedtuple of the form('name, prereqs, credits')
                prereqs is a tuple of prereqs where each prereq has the same form as the keys
        """
        wb = load_workbook("data/CourseCatalogSpring2020.xlsx")
        catalog = wb.get_sheet_by_name('catalog')
        Course = namedtuple('Course', 'program, designation')
        CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
        course_dict = {}
        for row in range(1, catalog.max_row + 1):
            key = Course(self.get_val(catalog, 'A', row), self.get_val(catalog, 'B', row))
            prereqs = tuple(tuple(self.get_split_course(prereq) for prereq in prereqs.split())
                       for prereqs in self.none_split(self.get_val(catalog, 'E', row)))
            val = CourseInfo(self.get_val(catalog, 'C', row), tuple(self.get_val(catalog, 'D', row).split()), prereqs)
            course_dict[key] = val
        return course_dict


    def get_split_course(self, course):
        """
        Parses a course from programdesignation into the ('program, designation') form.
        e.g. 'CS1101' -> ('CS', '1101')
        """
        return tuple(split_course for course_part in re.findall('((?:[A-Z]+-)?[A-Z]+)(.+)', course)
                     for split_course in course_part)


    def none_split(self, val):
        """Handles calling split on a None value by returning the empty list."""
        return val.split(', ') if val else ()


    def get_val(self, catalog, col, row):
        """Returns the value of a cell."""
        return catalog[col + str(row)].value


    def print_dict(self, dict):
        """Simply prints a dictionary's key and values line by line."""
        for key in dict:
            print(key, dict[key])
