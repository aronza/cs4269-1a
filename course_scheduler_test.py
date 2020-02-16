"""
    Team Number: 10
    Members: Noah Popham, Arda Turkmen, Mark Weinstein, Harry Wilson
"""
import sys
import os
import re
from collections import namedtuple
from pprint import pprint
import warnings

from openpyxl import load_workbook

from course_scheduler import course_scheduler

Course = namedtuple('Course', 'program, designation')
CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')


def create_course_dict(dictionary_file_path="data/CourseCatalogSpring2020.xlsx"):
    """
    Creates a dictionary containing course info.
    Keys: namedtuple of the form ('program, designation')
    Values: namedtuple of the form('name, prereqs, credits')
            prereqs is a tuple of prereqs where each prereq has the same form as the keys
    """
    wb = load_workbook(dictionary_file_path)
    catalog = wb.get_sheet_by_name('catalog')

    course_dict = {}
    for row in range(1, catalog.max_row + 1):
        key = Course(get_val(catalog, 'A', row), get_val(catalog, 'B', row))
        prereqs = tuple(tuple(get_split_course(prereq) for prereq in prereqs.split())
                        for prereqs in none_split(get_val(catalog, 'E', row)))
        val = CourseInfo(get_val(catalog, 'C', row), tuple(get_val(catalog, 'D', row).split()), prereqs)
        course_dict[key] = val
    return course_dict


def get_split_course(course):
    """
    Parses a course from programdesignation into the ('program, designation') form.
    e.g. 'CS1101' -> ('CS', '1101')
    """
    return tuple(split_course for course_part in re.findall('((?:[A-Z]+-)?[A-Z]+)(.+)', course)
                 for split_course in course_part)


def none_split(val):
    """Handles calling split on a None value by returning the empty list."""
    return val.split(', ') if val else ()


def get_val(catalog, col, row):
    """Returns the value of a cell."""
    return catalog[col + str(row)].value


def print_dict(dict):
    """Simply prints a dictionary's key and values line by line."""
    for key in dict:
        print(key, dict[key])


def test_with_catalog(catalog_file_path, test_type):
    """
    Tests the course_scheduler with a catalog in the given file path and the type of test that needs to be applied

    :param catalog_file_path: Absolute path to the excel file holding course catalog data.
    :param test_type: Current options are empty, have_all and have_subset.
        Empty expects an empty schedule because there is no viable schedule to be completed in 4 years.
        Have_all expects the returned schedule to include all the courses in the catalog.
        Have_subset expects the returned schedule to include a subset of the courses in the catalog.
    :return: True if test passes, else if it doesn't.
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        catalog = create_course_dict(catalog_file_path)
        validate_catalog(catalog)

        schedule = course_scheduler(course_descriptions=catalog, goal_conditions=[('CS', 'major')], initial_state=[])
        courses_scheduled = set(map(lambda scheduled_course: scheduled_course[0], schedule))

        if test_type == "empty":
            return courses_scheduled == set()
        if test_type == "have_all":
            return catalog.keys() == courses_scheduled
        if test_type == "have_subset":
            expectedPrograms = ['A', 'B', 'C', 'CS']
            expectedCourses = set(filter(lambda course: course.program in expectedPrograms, catalog.keys()))
            return expectedCourses.issubset(courses_scheduled)
        raise NameError(test_type + " is not a valid test option")


def validate_catalog(catalog):
    """
    Given code to test if the catalog makes sense.

    :param catalog: Python dictionary holding the catalog
    """
    # Test to see if all prereqs are in the file.
    prereq_list = [single_course for vals in catalog.values()
                   for some_prereqs in vals.prereqs for single_course in some_prereqs]
    for prereq in prereq_list:
        if prereq not in catalog:
            print(prereq)
    for key in catalog:
        # Test to see if every course has a term and credits.
        if not catalog[key].terms or not catalog[key].credits:
            print(key)
        # Test to see if a course's prereqs include the course itself
        if key in [course for prereq in catalog[key].prereqs for course in prereq]:
            print(key)


def main():
    """
    Main function that creates the default dictionary and runs the course_scheduler with ('CS', 'major') as goal
    and an empty initial State.
    """
    course_catalog = create_course_dict()
    validate_catalog(course_catalog)

    goal = [('CS', 'major')]
    courses_taken = []
    print("First search: ")
    print("Goal: " + str(goal))
    print("Initial State: " + str(courses_taken))
    print()
    schedule = course_scheduler(course_descriptions=course_catalog, goal_conditions=goal, initial_state=courses_taken)
    print()
    print("Result:")
    if type(schedule) is list:
        pprint(schedule)
    else:
        print(schedule)


if __name__ == "__main__":
    """
        Driver function that parses the command-line argument and calls the necessary functions.
        If the program is in test mode, call the test_with_catalog function on every catalog file in data/test_cases
    """
    TEST = (sys.argv[1] == "test") if len(sys.argv) > 1 else False

    if TEST:
        test_folder = os.path.join(os.path.dirname(__file__), 'data/test_cases/')
        tests_failed = 0
        test_count = 0

        for test_filename in os.listdir(test_folder):
            file_path = os.path.join(test_folder, test_filename)

            expected = None
            if test_filename == "test_no_solution.xlsx":
                expected = "empty"
            elif test_filename == "test_extra_classes.xlsx":
                expected = "have_subset"
            else:
                expected = "have_all"

            print("Test " + expected + " " + file_path)
            test_result = test_with_catalog(file_path, expected)
            print("\tPASS" if test_result else "\tFAIL")

            if not test_result:
                tests_failed += 1
            test_count += 1

        print(tests_failed, " Tests Failed. Number of tests run: ", test_count)
    else:
        main()
